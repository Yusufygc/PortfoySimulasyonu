# src/application/services/excel_export_service.py

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import List, Optional, Dict, Iterable, Union
import shutil

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.models.daily_price import DailyPrice
from src.domain.repositories.portfolio_repository import IPortfolioRepository
from src.domain.repositories.price_repository import IPriceRepository
from src.domain.repositories.stock_repository import IStockRepository
from src.domain.services_interfaces.i_market_data_client import IMarketDataClient


class ExportMode(str, Enum):
    OVERWRITE = "overwrite"
    APPEND = "append"


@dataclass
class DailyPosition:
    date: date
    ticker: str
    quantity: int
    avg_cost: Decimal
    cost_basis: Decimal
    close_price: Optional[Decimal]
    position_value: Optional[Decimal]
    daily_price_change_pct: Optional[Decimal]
    daily_pnl_tl: Optional[Decimal]          # <--- YENİ EKLENDİ
    unrealized_pnl_tl: Optional[Decimal]
    unrealized_pnl_pct: Optional[Decimal]
    weight_pct: Optional[Decimal]


@dataclass
class DailyPortfolioSnapshot:
    total_cost_basis: Optional[Decimal]
    date: date
    total_value: Optional[Decimal]
    daily_return_pct: Optional[Decimal]
    cumulative_return_pct: Optional[Decimal]
    daily_pnl: Optional[Decimal]
    cumulative_pnl: Optional[Decimal]
    status: str
    #active_stock_count: int 


class ExcelExportService:
    def __init__(
        self,
        portfolio_repo: IPortfolioRepository,
        price_repo: IPriceRepository,
        stock_repo: IStockRepository,
        market_data_client: IMarketDataClient,
    ) -> None:
        self.portfolio_repo = portfolio_repo
        self.price_repo = price_repo
        self.stock_repo = stock_repo
        self.market_data_client = market_data_client

    def _format_pct(self, value: Optional[Decimal]) -> Optional[float]:
        """
        Excel'e ham oranı (örn: 0.12) verir.
        100 ile çarpma işlemi YAPILMAZ.
        Formatlama Excel'in 'Percentage' stili ile (0.00%) halledilir.
        """
        if value is None:
            return None
        return float(value)

    def export_history(
        self,
        start_date: date,
        end_date: date,
        file_path: Union[str, Path],
        mode: ExportMode = ExportMode.OVERWRITE,
    ) -> None:
        file_path = Path(file_path)
        
        daily_positions, daily_snapshots = self._build_daily_history(start_date, end_date)
        
        detail_df = self._build_detail_df(daily_positions, daily_snapshots)
        summary_df = self._build_summary_df(daily_snapshots)
        stock_summary_df = self._build_stock_summary_df(daily_positions)
        dashboard_df = self._build_dashboard_df(daily_snapshots, daily_positions)

        if not file_path.exists():
            self._write_fresh_excel(file_path, summary_df, detail_df, stock_summary_df, dashboard_df)
            return

        if mode == ExportMode.OVERWRITE:
            self._write_fresh_excel(file_path, summary_df, detail_df, stock_summary_df, dashboard_df)
        else:
            self._append_to_existing_excel(file_path, summary_df, detail_df, stock_summary_df, dashboard_df)

    def _build_daily_history(self, start_date: date, end_date: date):
        if end_date < start_date:
            start_date, end_date = end_date, start_date

        all_trades = self.portfolio_repo.get_all_trades()
        relevant_trades = [t for t in all_trades if t.trade_date <= end_date]

        if not relevant_trades:
            return [], []

        from datetime import time as dt_time
        relevant_trades.sort(key=lambda t: (t.trade_date, getattr(t, "trade_time", None) or dt_time.min))

        stocks = self.stock_repo.get_all_stocks()
        ticker_map = {s.id: s.ticker for s in stocks}

        positions_state = {}
        trade_idx = 0
        n_trades = len(relevant_trades)

        daily_positions = []
        daily_snapshots = []
        last_close_by_stock = {}
        last_portfolio_value = None
        base_portfolio_value = None

        cur = start_date
        while cur <= end_date:
            while trade_idx < n_trades and relevant_trades[trade_idx].trade_date <= cur:
                t = relevant_trades[trade_idx]
                stock_id = t.stock_id
                qty = int(t.quantity)
                price = Decimal(str(t.price))

                if stock_id not in positions_state:
                    positions_state[stock_id] = {"qty": Decimal("0"), "total_cost": Decimal("0")}

                state = positions_state[stock_id]
                cur_qty = state["qty"]
                
                if t.side == "BUY":
                    state["qty"] += Decimal(qty)
                    state["total_cost"] += (Decimal(qty) * price)
                elif t.side == "SELL":
                    avg_cost = (state["total_cost"] / cur_qty) if cur_qty != 0 else Decimal("0")
                    sell_qty = min(Decimal(qty), cur_qty)
                    state["qty"] -= sell_qty
                    state["total_cost"] -= (avg_cost * sell_qty)
                    if state["qty"] == 0:
                        state["total_cost"] = Decimal("0")
                
                trade_idx += 1

            active_stock_ids = [sid for sid, st in positions_state.items() if st["qty"] > 0]
            prices_for_day = self.price_repo.get_prices_for_date(cur)
            
            is_weekend = cur.weekday() >= 5
            if not is_weekend and active_stock_ids:
                missing_ids = [sid for sid in active_stock_ids if sid not in prices_for_day]
                if missing_ids:
                    missing_tickers = [ticker_map.get(sid) for sid in missing_ids if sid in ticker_map]
                    if missing_tickers:
                        try:
                            fetched_prices = self.market_data_client.get_closing_prices(
                                stock_ids=missing_ids, tickers=missing_tickers, price_date=cur
                            )
                            new_daily_prices = []
                            for sid, price in fetched_prices.items():
                                prices_for_day[sid] = price
                                new_daily_prices.append(DailyPrice(
                                    id=None, stock_id=sid, price_date=cur, close_price=price
                                ))
                            if new_daily_prices:
                                self.price_repo.upsert_daily_prices_bulk(new_daily_prices)
                        except Exception:
                            pass

            has_prices = bool(prices_for_day)
            day_positions_list = []
            portfolio_value = None
            total_cost_basis = Decimal("0") 

            if has_prices:
                total_value = Decimal("0")
                temp_positions = []

                for stock_id, state in positions_state.items():
                    qty = int(state["qty"])
                    if qty <= 0: continue

                    avg_cost = (state["total_cost"] / state["qty"]) if state["qty"] != 0 else Decimal("0")
                    cost_basis = avg_cost * Decimal(qty)
                    total_cost_basis += cost_basis
                    
                    close_price = prices_for_day.get(stock_id)

                    if close_price:
                        pos_val = Decimal(qty) * close_price
                        total_value += pos_val
                        
                        last_c = last_close_by_stock.get(stock_id)
                        daily_chg = ((close_price / last_c) - 1) if (last_c and last_c != 0) else None
                        
                        # ÖNCE unrealized_tl hesapla
                        unrealized_tl = (close_price - avg_cost) * Decimal(qty)
                        unrealized_pct = (unrealized_tl / cost_basis) if cost_basis != 0 else None

                        # SONRA daily_pnl_stock hesapla
                        if last_c and last_c != 0:
                            # Normal gün: (Bugün - Dün) * Adet
                            daily_pnl_stock = (close_price - last_c) * Decimal(qty)
                        else:
                            # İlk gün: Toplam kar/zarar = günlük kar/zarar
                            daily_pnl_stock = unrealized_tl
                    else:
                        pos_val, daily_chg, daily_pnl_stock, unrealized_tl, unrealized_pct = None, None, None, None, None

                    temp_positions.append({
                        "ticker": ticker_map.get(stock_id, f"ID_{stock_id}"),
                        "qty": qty,
                        "avg_cost": avg_cost,
                        "cost_basis": cost_basis,
                        "close_price": close_price,
                        "pos_val": pos_val,
                        "daily_chg": daily_chg,
                        "daily_pnl": daily_pnl_stock,  # DOĞRU değişken adı
                        "unr_tl": unrealized_tl,
                        "unr_pct": unrealized_pct
                    })

                for tmp in temp_positions:
                    w_pct = (tmp["pos_val"] / total_value) if (tmp["pos_val"] and total_value) else None
                    day_positions_list.append(DailyPosition(
                        date=cur,
                        ticker=tmp["ticker"],
                        quantity=tmp["qty"],
                        avg_cost=tmp["avg_cost"],
                        cost_basis=tmp["cost_basis"],
                        close_price=tmp["close_price"],
                        position_value=tmp["pos_val"],
                        daily_price_change_pct=tmp["daily_chg"],
                        daily_pnl_tl=tmp["daily_pnl"],  # DOĞRU alan adı
                        unrealized_pnl_tl=tmp["unr_tl"],
                        unrealized_pnl_pct=tmp["unr_pct"],
                        weight_pct=w_pct
                    ))
                
                if total_value > 0:
                    portfolio_value = total_value
                    last_close_by_stock = prices_for_day.copy()
            else:
                portfolio_value = last_portfolio_value if last_portfolio_value else None

            daily_pnl, daily_ret, cum_pnl, cum_ret = None, None, None, None
            if portfolio_value:
                if not base_portfolio_value: 
                    base_portfolio_value = portfolio_value
                
                if last_portfolio_value:
                    # Normal gün
                    daily_pnl = portfolio_value - last_portfolio_value
                    daily_ret = (daily_pnl / last_portfolio_value)
                elif total_cost_basis > 0:
                    # İLK GÜN
                    daily_pnl = portfolio_value - total_cost_basis
                    daily_ret = (daily_pnl / total_cost_basis)
                
                # Kümülatif hesaplar
                if base_portfolio_value:
                    if portfolio_value == base_portfolio_value and total_cost_basis > 0:
                        cum_pnl = portfolio_value - total_cost_basis
                        cum_ret = (cum_pnl / total_cost_basis)
                    else:
                        if total_cost_basis > 0:
                            cum_pnl = portfolio_value - total_cost_basis
                            cum_ret = (cum_pnl / total_cost_basis)
                        else:
                            cum_pnl = portfolio_value - base_portfolio_value
                            cum_ret = (cum_pnl / base_portfolio_value)

            status = "Piyasa Açık" if has_prices else ("Hafta Sonu" if is_weekend else "Veri Yok")
            
            if day_positions_list:
                daily_positions.extend(day_positions_list)
            
            daily_snapshots.append(DailyPortfolioSnapshot(
                total_cost_basis=total_cost_basis,  # <--- FIXED
                date=cur, total_value=portfolio_value, 
                daily_return_pct=daily_ret, cumulative_return_pct=cum_ret,
                daily_pnl=daily_pnl, cumulative_pnl=cum_pnl, status=status
            ))

            if portfolio_value: last_portfolio_value = portfolio_value
            cur += timedelta(days=1)

        return daily_positions, daily_snapshots

    def _build_dashboard_df(self, snapshots: List[DailyPortfolioSnapshot], positions: List[DailyPosition]) -> pd.DataFrame:
        """Özet gösterge paneli"""
        if not snapshots:
            return pd.DataFrame()
        
        latest_snapshot = snapshots[-1]
        returns = [s.daily_return_pct for s in snapshots if s.daily_return_pct is not None]
        
        latest_date = latest_snapshot.date
        latest_positions = {}
        for p in positions:
            if p.date == latest_date:  # <--- Sadece son günü al
                latest_positions[p.ticker] = p
        
        best_stock = max(latest_positions.values(), 
                        key=lambda p: p.unrealized_pnl_pct if p.unrealized_pnl_pct else Decimal('-inf'))
        worst_stock = min(latest_positions.values(),
                         key=lambda p: p.unrealized_pnl_pct if p.unrealized_pnl_pct else Decimal('inf'))
        
        volatility = None
        if len(returns) > 1:
            returns_series = pd.Series([float(r) for r in returns])
            volatility = float(returns_series.std() * 100)
        
        max_drawdown = self._calculate_max_drawdown(snapshots)
        
        records = [
            {"Metrik": "Toplam Maliyet", "Değer": self._fmt_tr_money(latest_snapshot.total_cost_basis)},
            {"Metrik": "Güncel Portföy Değeri", "Değer": self._fmt_tr_money(latest_snapshot.total_value)},
            {"Metrik": "Toplam Kâr/Zarar (TL)", "Değer": self._fmt_tr_money(latest_snapshot.cumulative_pnl)},
            {"Metrik": "Toplam Getiri (%)", "Değer": self._fmt_tr_pct(latest_snapshot.cumulative_return_pct)},
            
            {"Metrik": "En İyi Performans", "Değer": f"{best_stock.ticker} ({float(best_stock.unrealized_pnl_pct * 100):.2f}%)" if best_stock.unrealized_pnl_pct else "N/A"},
            {"Metrik": "En Kötü Performans", "Değer": f"{worst_stock.ticker} ({float(worst_stock.unrealized_pnl_pct * 100):.2f}%)" if worst_stock.unrealized_pnl_pct else "N/A"},
            
        ]
        
        return pd.DataFrame(records)

    def _fmt_tr_money(self, val: Optional[Decimal]) -> str:
        if val is None: return "0,00"
        # 1234.56 -> 1.234,56
        s = f"{val:,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")

    def _fmt_tr_pct(self, val: Optional[Decimal]) -> str:
        if val is None: return "%0,00"
        # 0.1234 -> %12,34
        s = f"{val * 100:.2f}"
        return f"%{s.replace('.', ',')}"

    def _calculate_max_drawdown(self, snapshots: List[DailyPortfolioSnapshot]) -> Optional[float]:
        values = [float(s.total_value) for s in snapshots if s.total_value]
        if len(values) < 2:
            return None
        
        peak = values[0]
        max_dd = 0
        
        for value in values:
            if value > peak:
                peak = value
            dd = (peak - value) / peak * 100
            if dd > max_dd:
                max_dd = dd
        
        return max_dd

    def _build_detail_df(self, positions: Iterable[DailyPosition], snapshots: Iterable[DailyPortfolioSnapshot]) -> pd.DataFrame:
        snapshot_map = {s.date: s for s in snapshots}
        positions_by_date = {}
        for p in positions:
            if p.date not in positions_by_date:
                positions_by_date[p.date] = []
            positions_by_date[p.date].append(p)
            
        records = []
        sorted_dates = sorted(positions_by_date.keys())
        
        for d in sorted_dates:
            day_positions = positions_by_date[d]
            day_positions.sort(key=lambda x: x.ticker)
            
            total_cost_basis = Decimal("0")
            total_position_value = Decimal("0")
            total_unrealized_pnl_tl = Decimal("0")
            
            for p in day_positions:
                records.append({
                    "Tarih": p.date,
                    "Hisse": p.ticker,
                    "Adet": p.quantity,
                    "Ort. Maliyet (TL)": float(p.avg_cost),
                    "Güncel Fiyat (TL)": float(p.close_price) if p.close_price else None,
                    "Toplam Maliyet (TL)": float(p.cost_basis),
                    "Pozisyon Değeri (TL)": float(p.position_value) if p.position_value else None,
                    "Günlük Fiyat Değ. (%)": self._format_pct(p.daily_price_change_pct),
                    "Günlük K/Z (TL)": float(p.daily_pnl_tl) if p.daily_pnl_tl is not None else None, # <--- Eklendi
                    "Toplam K/Z (TL)": float(p.unrealized_pnl_tl) if p.unrealized_pnl_tl else None,
                    "Toplam K/Z (%)": self._format_pct(p.unrealized_pnl_pct),
                    "Portföy Ağırlığı (%)": self._format_pct(p.weight_pct),
                })
                
                if p.cost_basis: total_cost_basis += p.cost_basis
                if p.position_value: total_position_value += p.position_value
                if p.unrealized_pnl_tl: total_unrealized_pnl_tl += p.unrealized_pnl_tl

            snapshot = snapshot_map.get(d)
            total_unrealized_ratio = (total_unrealized_pnl_tl / total_cost_basis) if (total_cost_basis and total_cost_basis != 0) else None
            portfolio_daily_ret = snapshot.daily_return_pct if snapshot else None
            portfolio_daily_pnl = snapshot.daily_pnl if snapshot else None

            summary_row = {
                "Tarih":None,# d,
                "Hisse": "GÜNLÜK TOPLAM ➤➤➤",
                "Adet": None,
                "Ort. Maliyet (TL)": None,
                "Güncel Fiyat (TL)": None,
                "Toplam Maliyet (TL)": float(total_cost_basis),
                "Pozisyon Değeri (TL)": float(total_position_value),
                "Günlük Fiyat Değ. (%)": self._format_pct(portfolio_daily_ret),
                "Günlük K/Z (TL)": float(portfolio_daily_pnl) if portfolio_daily_pnl is not None else None,
                "Toplam K/Z (TL)": float(total_unrealized_pnl_tl),
                "Toplam K/Z (%)": self._format_pct(total_unrealized_ratio),
                "Portföy Ağırlığı (%)": self._format_pct(Decimal("1.0")),
            }
            records.append(summary_row)

        df = pd.DataFrame.from_records(records)
        return df

    def _build_stock_summary_df(self, positions: List[DailyPosition]) -> pd.DataFrame:
        if not positions:
            return pd.DataFrame()

        latest_positions = {}
        stock_days_count = {}

        for p in positions:
            latest_positions[p.ticker] = p
            stock_days_count[p.ticker] = stock_days_count.get(p.ticker, 0) + 1

        records = []
        for ticker, p in latest_positions.items():
            records.append({
                "Hisse": ticker,
                "Son Adet": p.quantity,
                "Ort. Maliyet (TL)": float(p.avg_cost),
                "Son Fiyat (TL)": float(p.close_price) if p.close_price else None,
                "Son Pozisyon Değeri (TL)": float(p.position_value) if p.position_value else None,
                "K/Z (TL)": float(p.unrealized_pnl_tl) if p.unrealized_pnl_tl else None,
                "K/Z (%)": self._format_pct(p.unrealized_pnl_pct),
                "Toplam Gün Sayısı": stock_days_count[ticker]
            })
        
        df = pd.DataFrame.from_records(records)
        if not df.empty:
            df = df.sort_values("Hisse").reset_index(drop=True)
        return df

    def _build_summary_df(self, snapshots: Iterable[DailyPortfolioSnapshot]) -> pd.DataFrame:
        records = []
        for s in snapshots:
            # Hafta sonlarını (veya veri olmayan tatil günlerini) atla
            if s.status == "Hafta Sonu":
                continue

            records.append({
                "Tarih": s.date,
                "Portföy Değeri (TL)": float(s.total_value) if s.total_value else None,
                "Günlük Getiri (%)": self._format_pct(s.daily_return_pct),
                "Toplam Getiri (%)": self._format_pct(s.cumulative_return_pct),
                "Günlük K/Z (TL)": float(s.daily_pnl) if s.daily_pnl else None,
                "Toplam K/Z (TL)": float(s.cumulative_pnl) if s.cumulative_pnl else None,
            })
        df = pd.DataFrame.from_records(records)
        if not df.empty:
            df = df.sort_values("Tarih").reset_index(drop=True)
        return df

    def _write_fresh_excel(self, file_path: Path, summary_df: pd.DataFrame, detail_df: pd.DataFrame, 
                          stock_summary_df: pd.DataFrame, dashboard_df: pd.DataFrame) -> None:
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                dashboard_df.to_excel(writer, sheet_name="Özet Panel", index=False)
                summary_df.to_excel(writer, sheet_name="Portföy Özeti", index=False)
                detail_df.to_excel(writer, sheet_name="Günlük Detaylar", index=False)
                stock_summary_df.to_excel(writer, sheet_name="Hisse Özeti", index=False)
                
                self._apply_formatting(writer, "Özet Panel", dashboard_df)
                self._apply_formatting(writer, "Portföy Özeti", summary_df)
                self._apply_formatting(writer, "Günlük Detaylar", detail_df)
                self._apply_formatting(writer, "Hisse Özeti", stock_summary_df)
                
        except PermissionError:
            raise PermissionError(f"Dosyaya yazılamadı: {file_path}\nDosya açık olabilir. Lütfen kapatıp tekrar deneyin.")

    def _apply_formatting(self, writer, sheet_name: str, df: pd.DataFrame):
        """Profesyonel Excel formatlaması"""
        if df.empty:
            return
            
        worksheet = writer.sheets[sheet_name]
        
        # 1. BAŞLIK SATIRI
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # 2. ZEBRASI SATIRLAR
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if row_num % 2 == 0:
                    cell.fill = light_gray
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 3. SAYISAL VE TARİH FORMATLARI
        for col_num, col_name in enumerate(df.columns, 1):
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Tarih Formatı (Mobil Uyumlu)
                if "Tarih" in col_name and cell.value is not None:
                    cell.number_format = 'dd.mm.yyyy'
                
                # TL sütunları
                elif "(TL)" in col_name and cell.value is not None:
                    if "Ort. Maliyet" in col_name or "Güncel Fiyat" in col_name or "Son Fiyat" in col_name:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0.00'
                
                # Yüzde sütunları
                elif "(%)" in col_name and cell.value is not None:
                    cell.number_format = '0.00%'
                
                # Adet sütunu
                elif col_name in ["Adet", "Son Adet", "Lot", "Toplam Gün Sayısı", "Aktif Pozisyon Sayısı"] and cell.value is not None:
                    cell.number_format = '#,##0'
        
        # 4. KOŞULLU RENKLENDIRME
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        
        for col_num, col_name in enumerate(df.columns, 1):
            # K/Z veya Getiri sütunları (Günlük K/Z dahil)
            if "K/Z" in col_name or "Getiri" in col_name:
                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif cell.value < 0:
                            cell.fill = red_fill
                            cell.font = red_font
        
        # 5. TOPLAM SATIRLARINI VURGULA
        bold_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_num in range(2, len(df) + 2):
            ticker_cell = None
            for col_num, col_name in enumerate(df.columns, 1):
                if "Hisse" in col_name or "Ticker" in col_name:
                    ticker_cell = worksheet.cell(row=row_num, column=col_num)
                    break
            
            if ticker_cell and ticker_cell.value and ("TOPLAM" in str(ticker_cell.value) or "▼" in str(ticker_cell.value)):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = bold_font
                    cell.fill = summary_fill
        
        # 6. SÜTUN GENİŞLİKLERİ
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            for row_num in range(2, min(len(df) + 2, 100)):
                cell_value = worksheet.cell(row=row_num, column=idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # 7. BAŞLIK SATIRINI DONDUR
        worksheet.freeze_panes = "A2"
        
        # 8. OTOMATİK FİLTRE
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

    def _append_to_existing_excel(self, file_path: Path, summary_df: pd.DataFrame, detail_df: pd.DataFrame,
                                  stock_summary_df: pd.DataFrame, dashboard_df: pd.DataFrame) -> None:
        try:
            with open(file_path, "r+"):
                pass
        except PermissionError:
            raise PermissionError(f"Dosya şu an açık: {file_path.name}\nLütfen Excel dosyasını kapatıp tekrar deneyin.")
        except Exception:
            pass

        try:
            with pd.ExcelFile(file_path, engine='openpyxl') as xls:
                sheet_names = xls.sheet_names
                
                existing_summary = pd.read_excel(xls, sheet_name="Portföy Özeti") if "Portföy Özeti" in sheet_names else pd.DataFrame()
                existing_detail = pd.read_excel(xls, sheet_name="Günlük Detaylar") if "Günlük Detaylar" in sheet_names else pd.DataFrame()
                existing_stock_sum = pd.read_excel(xls, sheet_name="Hisse Özeti") if "Hisse Özeti" in sheet_names else pd.DataFrame()
                existing_dashboard = pd.read_excel(xls, sheet_name="Özet Panel") if "Özet Panel" in sheet_names else pd.DataFrame()
        except Exception as e:
            print(f"Eski dosya okunamadı: {e}. Dosya yedeklenip yeniden oluşturulacak.")
            backup_path = file_path.with_suffix(file_path.suffix + ".bak")
            shutil.copy(file_path, backup_path)
            self._write_fresh_excel(file_path, summary_df, detail_df, stock_summary_df, dashboard_df)
            return

        combined_summary = pd.concat([existing_summary, summary_df], ignore_index=True)
        if not combined_summary.empty and "Tarih" in combined_summary.columns:
            combined_summary = combined_summary.drop_duplicates(subset=["Tarih"], keep="last").sort_values("Tarih").reset_index(drop=True)

        combined_detail = pd.concat([existing_detail, detail_df], ignore_index=True)
        if not combined_detail.empty and "Tarih" in combined_detail.columns and "Hisse" in combined_detail.columns:
            combined_detail = combined_detail.drop_duplicates(subset=["Tarih", "Hisse"], keep="last").sort_values(["Tarih", "Hisse"]).reset_index(drop=True)
            
        combined_stock_sum = pd.concat([existing_stock_sum, stock_summary_df], ignore_index=True)
        if not combined_stock_sum.empty and "Hisse" in combined_stock_sum.columns:
            combined_stock_sum = combined_stock_sum.drop_duplicates(subset=["Hisse"], keep="last").sort_values("Hisse").reset_index(drop=True)
        
        combined_dashboard = dashboard_df

        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                combined_dashboard.to_excel(writer, sheet_name="Özet Panel", index=False)
                combined_summary.to_excel(writer, sheet_name="Portföy Özeti", index=False)
                combined_detail.to_excel(writer, sheet_name="Günlük Detaylar", index=False)
                combined_stock_sum.to_excel(writer, sheet_name="Hisse Özeti", index=False)
                
                self._apply_formatting(writer, "Özet Panel", combined_dashboard)
                self._apply_formatting(writer, "Portföy Özeti", combined_summary)
                self._apply_formatting(writer, "Günlük Detaylar", combined_detail)
                self._apply_formatting(writer, "Hisse Özeti", combined_stock_sum)
        except PermissionError:
            raise PermissionError(f"Dosyaya yazılamadı: {file_path}\nDosya açık olabilir. Lütfen kapatıp tekrar deneyin.")