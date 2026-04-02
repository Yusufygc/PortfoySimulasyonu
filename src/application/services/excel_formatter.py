# src/application/services/excel_formatter.py

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    def apply_formatting(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """Profesyonel Excel formatlaması, durumsuz (stateless) operasyonlar."""
        if df.empty:
            return
            
        worksheet = writer.sheets[sheet_name]
        
        # 1. BAŞLIK SATIRI
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # 2. ZEBRASI SATIRLAR
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if row_num % 2 == 0:
                    cell.fill = light_gray
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 3. SAYISAL VE TARİH FORMATLARI
        for col_num, col_name in enumerate(df.columns, 1):
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                if "Tarih" in col_name and cell.value is not None:
                    cell.number_format = 'dd.mm.yyyy'
                elif "(TL)" in col_name and cell.value is not None:
                    if "Ort. Maliyet" in col_name or "Güncel Fiyat" in col_name or "Son Fiyat" in col_name:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0.00'
                elif "(%)" in col_name and cell.value is not None:
                    cell.number_format = '0.00%'
                elif col_name in ["Adet", "Son Adet", "Lot", "Toplam Gün Sayısı", "Aktif Pozisyon Sayısı"] and cell.value is not None:
                    cell.number_format = '#,##0'
        
        # 4. KOŞULLU RENKLENDIRME
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        
        for col_num, col_name in enumerate(df.columns, 1):
            if "K/Z" in col_name or "Getiri" in col_name:
                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif cell.value < 0:
                            cell.fill = red_fill
                            cell.font = red_font
        
        # 5. TOPLAM SATIRLARINI VURGULA
        bold_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_num in range(2, len(df) + 2):
            ticker_cell = None
            for col_num, col_name in enumerate(df.columns, 1):
                if "Hisse" in col_name or "Ticker" in col_name:
                    ticker_cell = worksheet.cell(row=row_num, column=col_num)
                    break
            
            if ticker_cell and ticker_cell.value and ("TOPLAM" in str(ticker_cell.value) or "▼" in str(ticker_cell.value)):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = bold_font
                    cell.fill = summary_fill
        
        # 6. SÜTUN GENİŞLİKLERİ
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            for row_num in range(2, min(len(df) + 2, 100)):
                cell_value = worksheet.cell(row=row_num, column=idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # 7. BAŞLIK SATIRINI DONDUR
        worksheet.freeze_panes = "A2"
        
        # 8. OTOMATİK FİLTRE
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
